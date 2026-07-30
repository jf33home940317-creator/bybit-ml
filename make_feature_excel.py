"""Generate feature summary Excel with all production + experimental features."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Feature Summary"

header_fill = PatternFill("solid", fgColor="1F4E79")
prod_fill = PatternFill("solid", fgColor="D6F5D6")
fail_fill = PatternFill("solid", fgColor="FFD6D6")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
normal_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", bold=True, size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 50

headers = ["#", "Feature", "Category", "Status", "Description"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

prod_features = [
    ("rsi_14", "Momentum", "RSI 14-period"),
    ("rsi_50", "Momentum", "RSI 50-period"),
    ("ppo", "Momentum", "Percentage Price Oscillator"),
    ("ppo_signal", "Momentum", "PPO Signal Line"),
    ("ppo_hist", "Momentum", "PPO Histogram"),
    ("natr_14", "Volatility", "Normalized ATR 14 (ATR/close)"),
    ("natr_72", "Volatility", "Normalized ATR 72 (ATR/close)"),
    ("bband_width_20", "Volatility", "Bollinger Band Width 20-period"),
    ("bband_width_50", "Volatility", "Bollinger Band Width 50-period"),
    ("ma_bias_20", "Trend", "(close - SMA20) / SMA20"),
    ("ma_bias_50", "Trend", "(close - SMA50) / SMA50"),
    ("ma_bias_200", "Trend", "(close - SMA200) / SMA200"),
    ("turnover_ratio_24", "Volume", "turnover / 24h rolling mean"),
    ("roc_4", "Momentum", "Rate of Change 4h"),
    ("roc_12", "Momentum", "Rate of Change 12h"),
    ("roc_24", "Momentum", "Rate of Change 24h"),
    ("hour_sin", "Time", "sin(2pi * hour / 24)"),
    ("hour_cos", "Time", "cos(2pi * hour / 24)"),
    ("dow_sin", "Time", "sin(2pi * dayofweek / 7)"),
    ("dow_cos", "Time", "cos(2pi * dayofweek / 7)"),
    ("is_weekend", "Time", "Saturday or Sunday = 1"),
    ("daily_rsi_14", "Daily", "Daily RSI 14 (shift +1d, no look-ahead)"),
    ("daily_ma_bias_20", "Daily", "Daily MA Bias 20 (shift +1d)"),
    ("daily_ma_bias_50", "Daily", "Daily MA Bias 50 (shift +1d)"),
    ("daily_ma_bias_200", "Daily", "Daily MA Bias 200 (shift +1d)"),
    ("daily_natr_14", "Daily", "Daily NATR 14 (shift +1d)"),
    ("cross_roc_24", "Cross-Asset", "BTC 24h ROC (ETH follows BTC with delay)"),
]

fail_8 = [
    ("return_1", "Momentum", "close.pct_change(1) -- redundant with roc_4"),
    ("return_3", "Momentum", "close.pct_change(3) -- redundant with roc_4"),
    ("return_6", "Momentum", "close.pct_change(6) -- redundant with roc_12"),
    ("ema20_50_ratio", "Trend", "EMA(20)/EMA(50) -- redundant with ma_bias"),
    ("volatility_24", "Volatility", "24h rolling std of returns -- redundant with natr"),
    ("obv_zscore", "Volume", "OBV z-score 200h -- noisy on 1h timeframe"),
    ("body_ratio", "Candlestick", "|close-open|/(high-low) -- noisy on 1h"),
    ("upper_shadow_ratio", "Candlestick", "(high-max(o,c))/(high-low) -- noisy on 1h"),
]

fail_fr_oi = [
    ("funding_rate", "Derivatives", "8h funding rate forward-filled -- too sparse for 1h model"),
    ("funding_rate_ma_24", "Derivatives", "FR 24h moving average"),
    ("funding_zscore_30d", "Derivatives", "FR z-score 30d -- unstable rolling std"),
    ("oi_change_1h", "Derivatives", "Open Interest 1h pct change"),
    ("oi_change_24h", "Derivatives", "Open Interest 24h pct change"),
    ("oi_price_divergence", "Derivatives", "sign(OI) != sign(price) -- collinear with price"),
]

row = 2

def write_feature(ws, row, idx, name, cat, status, desc, fill):
    ws.cell(row=row, column=1, value=idx).font = normal_font
    ws.cell(row=row, column=2, value=name).font = bold_font
    ws.cell(row=row, column=3, value=cat).font = normal_font
    sc = ws.cell(row=row, column=4, value=status)
    if status == "Production":
        sc.font = Font(name="Arial", bold=True, size=10, color="006400")
    else:
        sc.font = Font(name="Arial", bold=True, size=10, color="8B0000")
    sc.fill = fill
    ws.cell(row=row, column=5, value=desc).font = normal_font
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).alignment = Alignment(vertical="center")

for i, (name, cat, desc) in enumerate(prod_features, 1):
    write_feature(ws, row, i, name, cat, "Production", desc, prod_fill)
    row += 1

row += 1
sep = ws.cell(row=row, column=1, value="Experiment: +8 Technical (Sharpe 1.50 -> -0.45, REJECTED)")
sep.font = Font(name="Arial", bold=True, size=10, color="8B0000")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 1

for i, (name, cat, desc) in enumerate(fail_8, 28):
    write_feature(ws, row, i, name, cat, "Rejected", desc, fail_fill)
    row += 1

row += 1
sep = ws.cell(row=row, column=1, value="Experiment: Phase 6 FR/OI (Sharpe 1.50 -> 0.09, REJECTED)")
sep.font = Font(name="Arial", bold=True, size=10, color="8B0000")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 1

for i, (name, cat, desc) in enumerate(fail_fr_oi, 36):
    write_feature(ws, row, i, name, cat, "Rejected", desc, fail_fill)
    row += 1

# Sheet 2: Backtest Comparison
ws2 = wb.create_sheet("Backtest Comparison")
ws2.column_dimensions["A"].width = 25
for c in ["B", "C", "D"]:
    ws2.column_dimensions[c].width = 22

headers2 = ["Metric", "Original (27 feat)", "+8 Technical (35)", "+6 FR/OI (33)"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

data = [
    ["Sharpe Ratio", 1.5034, -0.4537, 0.0869],
    ["Total Return %", 98.27, -35.63, -1.19],
    ["Max Drawdown %", -13.26, -53.90, -19.77],
    ["Win Rate %", 54.76, 39.19, 54.76],
    ["Executed Trades", 84, 148, 50],
    ["Result", "PRODUCTION", "REJECTED", "REJECTED"],
]

for r, row_data in enumerate(data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = bold_font if c == 1 else normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
        if isinstance(val, str):
            if val == "PRODUCTION":
                cell.fill = prod_fill
                cell.font = Font(name="Arial", bold=True, size=10, color="006400")
            elif val == "REJECTED":
                cell.fill = fail_fill
                cell.font = Font(name="Arial", bold=True, size=10, color="8B0000")

out = "storage/excel/ETHUSDT_feature_summary.xlsx"
wb.save(out)
print(f"Saved: {out}")
